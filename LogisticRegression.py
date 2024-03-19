import openpyxl
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score

# Function to create a new Excel file
def create_excel(filename):
    workbook = openpyxl.Workbook()
    workbook.save(filename)
    print(f"Excel file '{filename}' created successfully.")

# Function to read data from Excel
def read_excel(filename):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    workbook.close()
    return data

# Function to write data to Excel
def write_excel(filename, data):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    for row in data:
        sheet.append(row)
    workbook.save(filename)
    workbook.close()
    print("Data written to Excel successfully.")

# Function to update data in Excel
def update_excel(filename, row_index, col_index, new_value):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    sheet.cell(row=row_index, column=col_index, value=new_value)
    workbook.save(filename)
    workbook.close()
    print("Data updated in Excel successfully.")

# Function to delete data from Excel
def delete_excel(filename, row_index):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    sheet.delete_rows(row_index)
    workbook.save(filename)
    workbook.close()
    print("Data deleted from Excel successfully.")

# Function to visualize data
def visualize_data(data):
    pass  # Since we're using Logistic Regression, no visualization is needed for this task

# Function to apply Logistic Regression
def logistic_regression(data):
    # Extracting features and labels
    X = np.array([row[1] for row in data[1:]])  # Age
    y = np.array([1 if row[2] == "New York" else 0 for row in data[1:]])  # Binary label: 1 if from New York, 0 otherwise

    # Splitting the data into training and testing sets
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

    # Training Logistic Regression model
    model = LogisticRegression()
    model.fit(X_train.reshape(-1, 1), y_train)

    # Making predictions
    y_pred = model.predict(X_test.reshape(-1, 1))

    # Calculating accuracy
    accuracy = accuracy_score(y_test, y_pred)
    print(f"Accuracy of Logistic Regression: {accuracy}")

# Example usage
filename = "example.xlsx"

# Create Excel file
create_excel(filename)

# Write data to Excel
data_to_write = [
    ["Name", "Age", "City"],
    ["Alice", 30, "New York"],
    ["Bob", 35, "Los Angeles"],
    ["Charlie", 25, "Chicago"],
    ["David", 40, "New York"],
    ["Eva", 28, "Los Angeles"],
    ["Frank", 32, "Chicago"],
    ["Grace", 45, "New York"],
    ["Henry", 27, "Los Angeles"],
    ["Isabella", 38, "Chicago"]
]
write_excel(filename, data_to_write)

# Read data from Excel
print("Data read from Excel:")
read_data = read_excel(filename)
for row in read_data:
    print(row)

# Update data in Excel
update_excel(filename, 3, 2, 28)

# Delete data from Excel
delete_excel(filename, 2)

# Read data from Excel after update and delete
print("Data read from Excel after update and delete:")
read_data = read_excel(filename)
for row in read_data:
    print(row)

# Apply Logistic Regression
logistic_regression(read_data)
