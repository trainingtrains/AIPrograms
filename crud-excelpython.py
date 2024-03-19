import openpyxl

# Function to create a new Excel file
def create_excel(filename):
    workbook = openpyxl.Workbook()
    workbook.save(filename)
    print(f"Excel file '{filename}' created successfully.")

# Function to read data from Excel
def read_excel(filename):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    workbook.close()
    return data

# Function to write data to Excel
def write_excel(filename, data):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    for row in data:
        sheet.append(row)
    workbook.save(filename)
    workbook.close()
    print("Data written to Excel successfully.")

# Function to update data in Excel
def update_excel(filename, row_index, col_index, new_value):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    sheet.cell(row=row_index, column=col_index, value=new_value)
    workbook.save(filename)
    workbook.close()
    print("Data updated in Excel successfully.")

# Function to delete data from Excel
def delete_excel(filename, row_index):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    sheet.delete_rows(row_index)
    workbook.save(filename)
    workbook.close()
    print("Data deleted from Excel successfully.")

# Example usage
filename = "mydata.xlsx"

# Create Excel file
create_excel(filename)

# Write data to Excel
data_to_write = [
    ["Name", "Age", "City"],
    ["Alice", 30, "New York"],
    ["Bob", 35, "Los Angeles"],
    ["Charlie", 25, "Chicago"],
    ["boopathi", 35, "erode"],
    ["kumar", 25, "erode"],
    
]
write_excel(filename, data_to_write)

# Read data from Excel
print("Data read from Excel:")
read_data = read_excel(filename)
for row in read_data:
    print(row)

# Update data in Excel
update_excel(filename, 3, 2, 28)

# Delete data from Excel
delete_excel(filename, 2)

# Read data from Excel after update and delete
print("Data read from Excel after update and delete:")
read_data = read_excel(filename)
for row in read_data:
    print(row)
