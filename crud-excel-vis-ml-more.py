import openpyxl
import matplotlib.pyplot as plt
import numpy as np
from sklearn.linear_model import LinearRegression

# Function to create a new Excel file
def create_excel(filename):
    workbook = openpyxl.Workbook()
    workbook.save(filename)
    print(f"Excel file '{filename}' created successfully.")

# Function to read data from Excel
def read_excel(filename):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    workbook.close()
    return data

# Function to write data to Excel
def write_excel(filename, data):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    for row in data:
        sheet.append(row)
    workbook.save(filename)
    workbook.close()
    print("Data written to Excel successfully.")

# Function to update data in Excel
def update_excel(filename, row_index, col_index, new_value):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    sheet.cell(row=row_index, column=col_index, value=new_value)
    workbook.save(filename)
    workbook.close()
    print("Data updated in Excel successfully.")

# Function to delete data from Excel
def delete_excel(filename, row_index):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    sheet.delete_rows(row_index)
    workbook.save(filename)
    workbook.close()
    print("Data deleted from Excel successfully.")

# Function to visualize data
def visualize_data(data):
    names = [row[0] for row in data[1:]]
    ages = [row[1] for row in data[1:]]
    cities = [row[2] for row in data[1:]]

    # Plotting age distribution
    plt.figure(figsize=(10, 6))
    plt.hist(ages, bins=np.arange(20, 60, 5), color='skyblue', edgecolor='black')
    plt.title('Age Distribution')
    plt.xlabel('Age')
    plt.ylabel('Frequency')
    plt.grid(True)
    plt.show()

    # Plotting city distribution
    unique_cities, city_counts = np.unique(cities, return_counts=True)
    plt.figure(figsize=(8, 6))
    plt.bar(unique_cities, city_counts, color='lightgreen')
    plt.title('City Distribution')
    plt.xlabel('City')
    plt.ylabel('Frequency')
    plt.xticks(rotation=45)
    plt.grid(True)
    plt.show()

# Function to apply machine learning algorithm
def apply_machine_learning(data):
    cities = np.array([row[2] for row in data[1:]]).reshape(-1, 1)
    ages = np.array([row[1] for row in data[1:]])

    # Convert city names to numerical labels
    unique_cities, city_labels = np.unique(cities, return_inverse=True)

    # Train linear regression model
    model = LinearRegression()
    model.fit(city_labels.reshape(-1, 1), ages)

    # Predict ages for all cities
    predicted_ages = model.predict(np.arange(len(unique_cities)).reshape(-1, 1))

    # Plotting predicted ages
    plt.figure(figsize=(8, 6))
    plt.bar(unique_cities, predicted_ages, color='orange')
    plt.title('Predicted Age based on City')
    plt.xlabel('City')
    plt.ylabel('Predicted Age')
    plt.xticks(rotation=45)
    plt.grid(True)
    plt.show()

# Example usage
filename = "example.xlsx"

# Create Excel file
create_excel(filename)

# Write data to Excel
data_to_write = [
    ["Name", "Age", "City"],
    ["Alice", 30, "New York"],
    ["Bob", 35, "Los Angeles"],
    ["Charlie", 25, "Chicago"],
    ["David", 40, "New York"],
    ["Eva", 28, "Los Angeles"],
    ["Frank", 32, "Chicago"],
    ["Grace", 45, "New York"],
    ["Henry", 27, "Los Angeles"],
    ["Isabella", 38, "Chicago"]
]
write_excel(filename, data_to_write)

# Read data from Excel
print("Data read from Excel:")
read_data = read_excel(filename)
for row in read_data:
    print(row)

# Update data in Excel
update_excel(filename, 3, 2, 28)

# Delete data from Excel
delete_excel(filename, 2)

# Read data from Excel after update and delete
print("Data read from Excel after update and delete:")
read_data = read_excel(filename)
for row in read_data:
    print(row)

# Visualize data
visualize_data(read_data)

# Apply machine learning algorithm
apply_machine_learning(read_data)
